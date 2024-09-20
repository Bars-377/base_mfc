from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file
from models import db, Service, User
from datetime import datetime
import pandas as pd
from io import BytesIO
from openpyxl.styles import PatternFill, Border, Side
from flask_login import LoginManager, login_user, logout_user, login_required

app = Flask(__name__)

import os
app.secret_key = os.urandom(24)  # Генерирует случайный ключ длиной 24 байта

# app.secret_key = 'supersecretkey'

app.config['SQLALCHEMY_DATABASE_URI'] = 'mysql+pymysql://root:enigma1418@localhost/mdtomskbot'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db.init_app(app)

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# Изменим сообщение о необходимости входа
login_manager.login_message = 'Для доступа к этой странице, пожалуйста, войдите в систему.'
login_manager.login_message_category = 'info'  # Можно задать категорию, например, 'info', 'warning', 'danger'

@login_manager.user_loader
def load_user(user_id):
    return User.query.get_or_404(int(user_id))

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('index'))

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        new_user = User(username=username)
        new_user.set_password(password)  # Предполагаем, что метод set_password уже реализован
        db.session.add(new_user)
        db.session.commit()
        flash('Registration successful!', 'success')
        return redirect(url_for('login'))
    return render_template('register.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = User.query.filter_by(username=username).first()
        if user and user.check_password(password):
            login_user(user)
            flash('Вход успешен!', 'success')
            next_page = request.args.get('next')
            return redirect(next_page or url_for('index'))
        flash('Неверное имя пользователя или пароль!', 'danger')
    return render_template('login.html')

@app.route('/')
@login_required
def index():
    year = request.args.get('year', None)
    keyword = request.args.get('keyword', None)
    selected_column = request.args.get('column', None)
    page = request.args.get('page', 1, type=int)
    per_page = 20

    service_years = db.session.query(db.func.year(Service.year)).distinct().all()
    service_years = [str(year[0]) for year in service_years]

    query = Service.query

    if year == 'No':
        year = None

    if year == 'None':  # Если year == 'None', фильтруем записи, у которых год == NULL
        query = query.filter(Service.year.is_(None))
    elif year:
        query = query.filter(db.func.year(Service.year) == year)

    if keyword:
        if selected_column and hasattr(Service, selected_column):
            column = getattr(Service, selected_column)
            query = query.filter(column.like(f'%{keyword}%'))
        else:
            columns = [c for c in Service.__table__.columns.keys()]
            filters = [getattr(Service, col).like(f'%{keyword}%') for col in columns]
            query = query.filter(db.or_(*filters))

    """Сортировка в cast(Service.id_id, Integer).asc() идёт с преобразованием в Integer"""
    from sqlalchemy import cast, Integer
    query = query.order_by(cast(Service.id_id, Integer).asc(), Service.year.asc())

    total_cost_1 = db.session.query(db.func.sum(Service.cost)).scalar() or 0
    total_cost_2 = db.session.query(db.func.sum(Service.certificate)).scalar() or 0
    total_cost_3 = db.session.query(db.func.sum(Service.certificate_no)).scalar() or 0

    offset = (page - 1) * per_page
    services = query.offset(offset).limit(per_page).all()
    total_services = query.count()
    total_pages = (total_services + per_page - 1) // per_page

    max_page_buttons = 5
    start_page = max(1, page - max_page_buttons // 2)
    end_page = min(total_pages, page + max_page_buttons // 2)

    if end_page - start_page < max_page_buttons - 1:
        if start_page > 1:
            end_page = min(total_pages, end_page + (max_page_buttons - (end_page - start_page)))
        else:
            start_page = max(1, end_page - (max_page_buttons - (end_page - start_page)))

    return render_template(
        'index.html',
        services=services,
        total_cost_1=total_cost_1,
        total_cost_2=total_cost_2,
        total_cost_3=total_cost_3,
        selected_year=year,
        selected_column=selected_column,
        keyword=keyword,
        page=page,
        total_pages=total_pages,
        start_page=start_page,
        end_page=end_page,
        service_years=service_years
    )

@app.route('/edit/<int:id>', methods=['GET'])
@login_required
def edit(id):
    service = Service.query.get_or_404(id)
    return render_template('edit.html', service=service)

@app.route('/add_edit', methods=['GET'])
@login_required
def add_edit():
    return render_template('add.html')

@app.route('/edit/<int:id>', methods=['POST'])
@login_required
def update(id):
    service = Service.query.get_or_404(id)
    service.id_id = request.form['id_id']
    service.name = request.form['name']
    service.snils = request.form['snils']
    service.location = request.form['location']
    service.address_p = request.form['address_p']
    service.address = request.form['address']
    service.benefit = request.form['benefit']
    service.number = request.form['number']
    service.year = datetime.strptime(request.form['year'], '%d.%m.%Y').date()
    service.cost = request.form['cost']
    service.certificate = request.form['certificate']
    service.date_number_get = request.form['date_number_get']
    service.date_number_cancellation = request.form['date_number_cancellation']
    service.date_number_no = request.form['date_number_no']
    service.certificate_no = request.form['certificate_no']
    service.reason = request.form['reason']
    service.track = request.form['track']
    service.date_post = request.form['date_post']
    service.color = request.form.get('color')

    db.session.commit()
    flash('Данные успешно обновлены!', 'success')
    return redirect(url_for('index'))

@app.route('/update-color/<int:id>', methods=['POST'])
@login_required
def update_color(id):
    service = Service.query.get_or_404(id)
    data = request.json
    color = data.get('color')

    if color:
        service.color = color
        db.session.commit()
        return jsonify({
            'success': True,
            'id': service.id,
            'color': service.color
        })
    else:
        return jsonify({'success': False, 'message': 'No color provided'}), 400

@app.route('/delete/<int:id>', methods=['POST'])
@login_required
def delete(id):
    service = Service.query.get_or_404(id)
    db.session.delete(service)
    db.session.commit()
    flash('Данные успешно удалены!', 'success')
    return redirect(url_for('index'))

@app.route('/add', methods=['POST'])
@login_required
def add():
    id_id = request.form['id_id']
    name = request.form['name']
    snils = request.form['snils']
    location = request.form['location']
    address_p = request.form['address_p']
    address = request.form['address']
    benefit = request.form['benefit']
    number = request.form['number']
    year = datetime.strptime(request.form['year'], '%d.%m.%Y').date()
    cost = request.form['cost']
    certificate = request.form['certificate']
    date_number_get = request.form['date_number_get']
    date_number_cancellation = request.form['date_number_cancellation']
    date_number_no = request.form['date_number_no']
    certificate_no = request.form['certificate_no']
    reason = request.form['reason']
    track = request.form['track']
    date_post = request.form['date_post']
    color = request.form.get('color')

    new_service = Service(id_id=id_id, name=name, snils=snils, location=location,
                        address_p=address_p, address=address, benefit=benefit,
                        number=number, year=year, cost=cost,
                        certificate=certificate, date_number_get=date_number_get,
                        date_number_cancellation=date_number_cancellation,
                        date_number_no=date_number_no, certificate_no=certificate_no,
                        reason=reason, track=track, date_post=date_post, color=color)
    db.session.add(new_service)
    db.session.commit()
    flash('Данные успешно добавлены!', 'success')
    return redirect(url_for('index'))

@app.route('/export-excel', methods=['GET'])
@login_required
def export_excel():
    year = request.args.get('year', None)

    query = Service.query

    if year == 'None':  # Если year == 'None', фильтруем записи, у которых год == NULL
        query = query.filter(Service.year.is_(None))
    elif year:
        query = query.filter(db.func.year(Service.year) == year)

    services = query.all()

    df = pd.DataFrame([{
        '№ п/п': service.id_id,
        'ФИО заявителя': service.name,
        'СНИЛС': service.snils,
        'Район': service.location,
        'Населённый пункт': service.address_p,
        'Адрес': service.address,
        'Льгота': service.benefit,
        'Серия и номер': service.number,
        'Дата выдачи сертификата': service.year.strftime('%Y-%m-%d') if service.year else None,
        'Размер выплаты': service.cost,
        'Сертификат': service.certificate,
        'Дата и номер решения о выдаче': service.date_number_get,
        'Дата и № решения об аннулировании': service.date_number_cancellation,
        'Дата и № решения об отказе в выдаче': service.date_number_no,
        'Отказ в выдаче': service.certificate_no,
        'Причина отказа': service.reason,
        'ТРЕК': service.track,
        'Дата отправки почтой': service.date_post,
        'Color': getattr(service, 'color', '')
    } for service in services])

    # Приводим колонки к числовому типу данных
    df['Размер выплаты'] = pd.to_numeric(df['Размер выплаты'], errors='coerce')
    df['Сертификат'] = pd.to_numeric(df['Сертификат'], errors='coerce')
    df['Отказ в выдаче'] = pd.to_numeric(df['Отказ в выдаче'], errors='coerce')

    # Расчет итогов
    total_cost = df['Размер выплаты'].sum()
    total_certificate = df['Сертификат'].sum()
    total_certificate_no = df['Отказ в выдаче'].sum()

    # Создание строки с итогами
    totals_row = pd.DataFrame([{
        '№ п/п': '',
        'ФИО заявителя': '',
        'СНИЛС': '',
        'Район': '',
        'Населённый пункт': '',
        'Адрес': '',
        'Льгота': '',
        'Серия и номер': '',
        'Дата выдачи сертификата': '',
        'Размер выплаты': total_cost,
        'Сертификат': total_certificate,
        'Дата и номер решения о выдаче': '',
        'Дата и № решения об аннулировании': '',
        'Дата и № решения об отказе в выдаче': '',
        'Отказ в выдаче': total_certificate_no,
        'Причина отказа': '',
        'ТРЕК': '',
        'Дата отправки почтой': '',
        'Color': ''
    }])

    # Добавление строки с итогами в DataFrame
    df = pd.concat([df, totals_row], ignore_index=True)

    # Создаем файл Excel
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Services')

        # Настройка стиля
        worksheet = writer.sheets['Services']

        # Определяем стиль для границ
        border_style = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))

        # Определяем стиль для заливки желтым цветом
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        # Применяем границы ко всем ячейкам и цвет к ячейкам, где он задан
        for row_num in range(2, worksheet.max_row + 1):  # Пропускаем заголовки
            for col_num in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.border = border_style

                # Применяем цвет только к ячейкам данных
                color = df.iloc[row_num - 2]['Color']  # Сопоставление индексов DataFrame
                if color:
                    cell.fill = PatternFill(start_color=color.replace('#', ''), end_color=color.replace('#', ''), fill_type="solid")

        # Применяем границы к заголовкам
        for col_num in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.border = border_style

        # Применяем желтый цвет к строке с итогами
        totals_row_num = worksheet.max_row
        for col_num in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=totals_row_num, column=col_num)
            cell.fill = yellow_fill
            cell.border = border_style

        # Удаляем столбец Color из Excel-файла
        worksheet.delete_cols(df.columns.get_loc("Color") + 1)

    output.seek(0)

    response = send_file(output, as_attachment=True, download_name='services.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    # Отправляем файл пользователю
    return response

"""Nginx"""
from waitress import serve
if __name__ == '__main__':
    print('Flask для Nginx запущен!')
    serve(app, threads=10, host='172.18.11.103', port=8000)

# """Standart"""
# if __name__ == '__main__':
#     print('Flask запущен')
#     app.run(host='0.0.0.0', port=5000, debug=True)